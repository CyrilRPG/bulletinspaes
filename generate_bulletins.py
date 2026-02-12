#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generateur de Bulletins Scolaires - PAES + Linova Education
Lit les fichiers Excel du dossier notes/ (1 fichier = 1 matiere),
applique la formule de reajustement, genere des bulletins HTML/PDF
pour chaque eleve en version PAES et Linova.
"""

import os
import gc
from pathlib import Path
from openpyxl import load_workbook
from weasyprint import HTML
import unicodedata
import re
import random

# === CONFIGURATION ===

# Profil PAES (Diploma Sante)
PROFIL_PAES = {
    "nom": "Diploma Santé",
    "adresse": "85 Avenue Ledru Rollin",
    "code_postal": "75012",
    "ville": "Paris",
    "annee_scolaire": "2025/2026",
    "classe": "PAES",
    "charge_etudes": "Shirel Benchetrit",
    "semestre": "Annuel",
    "logo": "logo_etablissement.png",
    "tampon": "tampon.png",
}

# Profil Linova Education
PROFIL_LINOVA = {
    "nom": "Linova Education",
    "adresse": "100 Quai de la Rapée",
    "code_postal": "75012",
    "ville": "Paris",
    "annee_scolaire": "2025/2026",
    "classe": "1",
    "charge_etudes": "Cyril Robert",
    "semestre": "Annuel",
    "logo": "logo_linova.png",
    "tampon": "tampon_linova.jpeg",
}

# Matieres PAES avec mapping vers fichiers Excel
MATIERES_PAES = [
    {"nom": "Biochimie", "excel_file": "Resultats-CB1 - Biochimie (1).xlsx", "enseignant": "M. Benramdane"},
    {"nom": "Biologie Cellulaire", "excel_file": "Resultats-CB1 - Biologie Cellulaire  (1).xlsx", "enseignant": "M. Descatoire"},
    {"nom": "Biostatistiques", "excel_file": "Resultats-CB1 - Biostatistiques  (1).xlsx", "enseignant": "U. Bederede"},
    {"nom": "Chimie Médecine", "excel_file": "Resultats-CB1 - Chimie  (2).xlsx", "enseignant": "R. Hadjerci"},
    {"nom": "Chimie Terminale", "excel_file": "Resultats-CB1 - Chimie  (3).xlsx", "enseignant": "D. Yazidi"},
    {"nom": "Mathématiques", "excel_file": "Resultats-CB1 - Maths (1).xlsx", "enseignant": "U. Bederede"},
    {"nom": "Physique", "excel_file": "Resultats-CB1 - Physique (1).xlsx", "enseignant": "H. Diaw"},
    {"nom": "Physique/Biophysique", "excel_file": "Resultats-CB1 - Physique Biophysique  (1).xlsx", "enseignant": "H. Diaw"},
    {"nom": "SVT", "excel_file": "Resultats-CB1 - SVT (1).xlsx", "enseignant": "M. Descatoire"},
]

# Matieres Linova avec mapping vers les matieres PAES (pour reutiliser les notes)
MATIERES_LINOVA = [
    {"nom": "Démarche qualité et organisation opérationnelle au sein du laboratoire de biologie médicale",
     "source_paes": "Biochimie", "enseignant": "M. Benramdane"},
    {"nom": "Analyses médicales les plus courantes",
     "source_paes": "Biologie Cellulaire", "enseignant": "M. Descatoire"},
    {"nom": "Amélioration des méthodes d'analyse de biologie médicale - Pratiques à visée thérapeutique",
     "source_paes": "Biostatistiques", "enseignant": "M. Descatoire"},
    {"nom": "Relations, collaboration et développement professionnels",
     "source_paes": "Chimie Médecine", "enseignant": "M. Benramdane"},
    {"nom": "Prélèvements de sang et d'autres échantillons biologiques / Culture générale et expression",
     "source_paes": "Chimie Terminale", "enseignant": "Aurore Pottier"},
    {"nom": "Anglais",
     "source_paes": "Physique/Biophysique", "enseignant": "Hanna Charbit"},
    {"nom": "SVT",
     "source_paes": "SVT", "enseignant": "M. Descatoire"},
    {"nom": "Mathématiques",
     "source_paes": "Mathématiques", "enseignant": "MOKRANE Zahia"},
    {"nom": "Physique-Chimie",
     "source_paes": "Physique", "enseignant": "MOKRANE Zahia"},
]

# Appreciations par matiere PAES
APPRECIATIONS_PAES = {
    "Biochimie": {
        "insuffisant": "Les notions de biochimie nécessitent un approfondissement. Un travail régulier permettra de progresser.",
        "passable": "Les bases en biochimie sont acquises. Poursuivez vos efforts pour consolider vos connaissances.",
        "assez_bien": "Bonne compréhension des concepts biochimiques. Continuez sur cette voie prometteuse.",
        "bien": "Très bonne maîtrise de la biochimie. Vos efforts sont récompensés, persévérez.",
        "excellent": "Excellente maîtrise des notions biochimiques. Félicitations pour ce travail remarquable."
    },
    "Biologie Cellulaire": {
        "insuffisant": "Les mécanismes cellulaires doivent être revus. Un travail plus soutenu est nécessaire.",
        "passable": "Compréhension correcte de la biologie cellulaire. Des efforts supplémentaires consolideront vos acquis.",
        "assez_bien": "Bonne assimilation des concepts cellulaires. Maintenez cette dynamique positive.",
        "bien": "Très bonne compréhension des processus cellulaires. Continuez ainsi.",
        "excellent": "Maîtrise remarquable de la biologie cellulaire. Travail exemplaire."
    },
    "Biostatistiques": {
        "insuffisant": "Les méthodes statistiques demandent plus de pratique. Un entraînement régulier est conseillé.",
        "passable": "Les bases statistiques sont comprises. Continuez à vous exercer pour gagner en aisance.",
        "assez_bien": "Bonne application des outils statistiques. Poursuivez vos efforts.",
        "bien": "Très bonne maîtrise des biostatistiques. Résultats très satisfaisants.",
        "excellent": "Excellente compréhension et application des méthodes statistiques. Bravo."
    },
    "Chimie Médecine": {
        "insuffisant": "Les fondamentaux en chimie médicale doivent être renforcés. Travaillez régulièrement.",
        "passable": "Niveau correct en chimie médicale. Poursuivez vos efforts pour vous améliorer.",
        "assez_bien": "Bonne compréhension de la chimie appliquée à la médecine. Continuez ainsi.",
        "bien": "Très bon niveau en chimie médicale. Vos résultats sont encourageants.",
        "excellent": "Excellente maîtrise de la chimie médicale. Félicitations pour votre investissement."
    },
    "Chimie Terminale": {
        "insuffisant": "Les acquis de chimie terminale nécessitent une révision. Un travail soutenu s'impose.",
        "passable": "Les notions de chimie terminale sont assimilées. Continuez à progresser.",
        "assez_bien": "Bonne maîtrise des concepts de chimie terminale. Persévérez dans vos efforts.",
        "bien": "Très bonne compréhension de la chimie terminale. Résultats très positifs.",
        "excellent": "Excellents résultats en chimie terminale. Travail remarquable et rigoureux."
    },
    "Mathématiques": {
        "insuffisant": "Les compétences mathématiques doivent être consolidées. Un entraînement quotidien est recommandé.",
        "passable": "Niveau satisfaisant en mathématiques. Continuez à pratiquer pour progresser.",
        "assez_bien": "Bonne maîtrise des outils mathématiques. Maintenez vos efforts.",
        "bien": "Très bon niveau en mathématiques. Vos compétences sont solides.",
        "excellent": "Excellente maîtrise des mathématiques. Résultats impressionnants, félicitations."
    },
    "Physique": {
        "insuffisant": "Les concepts physiques nécessitent plus de travail. Revoyez les notions fondamentales.",
        "passable": "Compréhension correcte des phénomènes physiques. Poursuivez vos efforts.",
        "assez_bien": "Bonne assimilation des lois physiques. Continuez sur cette lancée positive.",
        "bien": "Très bonne maîtrise de la physique. Vos efforts portent leurs fruits.",
        "excellent": "Excellente compréhension de la physique. Travail exemplaire et rigoureux."
    },
    "Physique/Biophysique": {
        "insuffisant": "Les notions de biophysique demandent un approfondissement. Travaillez régulièrement.",
        "passable": "Les bases en biophysique sont acquises. Continuez à consolider vos connaissances.",
        "assez_bien": "Bonne compréhension des applications physiques en biologie. Persévérez.",
        "bien": "Très bon niveau en biophysique. Résultats très encourageants.",
        "excellent": "Maîtrise excellente de la biophysique. Félicitations pour ce parcours remarquable."
    },
    "SVT": {
        "insuffisant": "Les connaissances en SVT doivent être renforcées. Un travail plus régulier est nécessaire.",
        "passable": "Niveau correct en SVT. Poursuivez vos efforts pour améliorer vos résultats.",
        "assez_bien": "Bonne compréhension des sciences de la vie et de la Terre. Continuez ainsi.",
        "bien": "Très bonne maîtrise des SVT. Vos résultats reflètent un travail sérieux.",
        "excellent": "Excellents résultats en SVT. Travail remarquable et approfondi, bravo."
    }
}

# Appreciations par matiere Linova
APPRECIATIONS_LINOVA = {
    "Démarche qualité et organisation opérationnelle au sein du laboratoire de biologie médicale": {
        "insuffisant": "Les notions de démarche qualité en laboratoire nécessitent un approfondissement. Un travail régulier permettra de progresser.",
        "passable": "Les bases de la démarche qualité sont acquises. Poursuivez vos efforts pour consolider vos connaissances.",
        "assez_bien": "Bonne compréhension de l'organisation opérationnelle en laboratoire. Continuez sur cette voie prometteuse.",
        "bien": "Très bonne maîtrise de la démarche qualité et de l'organisation en laboratoire. Persévérez.",
        "excellent": "Excellente maîtrise de la démarche qualité en laboratoire. Félicitations pour ce travail remarquable."
    },
    "Analyses médicales les plus courantes": {
        "insuffisant": "Les techniques d'analyses médicales doivent être revues. Un travail plus soutenu est nécessaire.",
        "passable": "Compréhension correcte des analyses médicales courantes. Des efforts supplémentaires consolideront vos acquis.",
        "assez_bien": "Bonne assimilation des méthodes d'analyses médicales. Maintenez cette dynamique positive.",
        "bien": "Très bonne compréhension des analyses médicales courantes. Continuez ainsi.",
        "excellent": "Maîtrise remarquable des analyses médicales les plus courantes. Travail exemplaire."
    },
    "Amélioration des méthodes d'analyse de biologie médicale - Pratiques à visée thérapeutique": {
        "insuffisant": "Les méthodes d'amélioration des analyses demandent plus de pratique. Un entraînement régulier est conseillé.",
        "passable": "Les bases des méthodes d'analyse sont comprises. Continuez à vous exercer pour gagner en aisance.",
        "assez_bien": "Bonne application des méthodes d'analyse de biologie médicale. Poursuivez vos efforts.",
        "bien": "Très bonne maîtrise des méthodes d'analyse à visée thérapeutique. Résultats très satisfaisants.",
        "excellent": "Excellente compréhension et application des méthodes d'analyse de biologie médicale. Bravo."
    },
    "Relations, collaboration et développement professionnels": {
        "insuffisant": "Les compétences relationnelles et collaboratives doivent être renforcées. Travaillez régulièrement.",
        "passable": "Niveau correct en relations et collaboration professionnelles. Poursuivez vos efforts.",
        "assez_bien": "Bonne compréhension des enjeux de collaboration et développement professionnel. Continuez ainsi.",
        "bien": "Très bon niveau en relations et développement professionnels. Résultats encourageants.",
        "excellent": "Excellente maîtrise des compétences relationnelles et professionnelles. Félicitations."
    },
    "Prélèvements de sang et d'autres échantillons biologiques / Culture générale et expression": {
        "insuffisant": "Les techniques de prélèvement et les compétences en expression nécessitent une révision. Un travail soutenu s'impose.",
        "passable": "Les notions de prélèvement et de culture générale sont assimilées. Continuez à progresser.",
        "assez_bien": "Bonne maîtrise des techniques de prélèvement et de l'expression. Persévérez dans vos efforts.",
        "bien": "Très bonne compréhension des prélèvements biologiques et de la culture générale. Résultats très positifs.",
        "excellent": "Excellents résultats en prélèvements et culture générale. Travail remarquable et rigoureux."
    },
    "Anglais": {
        "insuffisant": "Les compétences en anglais doivent être consolidées. Un entraînement quotidien est recommandé.",
        "passable": "Niveau satisfaisant en anglais. Continuez à pratiquer pour progresser.",
        "assez_bien": "Bonne maîtrise de l'anglais. Maintenez vos efforts.",
        "bien": "Très bon niveau en anglais. Vos compétences linguistiques sont solides.",
        "excellent": "Excellente maîtrise de l'anglais. Résultats impressionnants, félicitations."
    },
    "SVT": {
        "insuffisant": "Les connaissances en SVT doivent être renforcées. Un travail plus régulier est nécessaire.",
        "passable": "Niveau correct en SVT. Poursuivez vos efforts pour améliorer vos résultats.",
        "assez_bien": "Bonne compréhension des sciences de la vie et de la Terre. Continuez ainsi.",
        "bien": "Très bonne maîtrise des SVT. Vos résultats reflètent un travail sérieux.",
        "excellent": "Excellents résultats en SVT. Travail remarquable et approfondi, bravo."
    },
    "Mathématiques": {
        "insuffisant": "Les compétences mathématiques doivent être consolidées. Un entraînement quotidien est recommandé.",
        "passable": "Niveau satisfaisant en mathématiques. Continuez à pratiquer pour progresser.",
        "assez_bien": "Bonne maîtrise des outils mathématiques. Maintenez vos efforts.",
        "bien": "Très bon niveau en mathématiques. Vos compétences sont solides.",
        "excellent": "Excellente maîtrise des mathématiques. Résultats impressionnants, félicitations."
    },
    "Physique-Chimie": {
        "insuffisant": "Les concepts de physique-chimie nécessitent plus de travail. Revoyez les notions fondamentales.",
        "passable": "Compréhension correcte de la physique-chimie. Poursuivez vos efforts.",
        "assez_bien": "Bonne assimilation des lois de physique-chimie. Continuez sur cette lancée positive.",
        "bien": "Très bonne maîtrise de la physique-chimie. Vos efforts portent leurs fruits.",
        "excellent": "Excellente compréhension de la physique-chimie. Travail exemplaire et rigoureux."
    }
}

# Chemins
BASE_DIR = Path(__file__).parent
NOTES_DIR = BASE_DIR / "notes"
TEMPLATE_FILE = BASE_DIR / "bulletin_template.html"
TEMPLATE_LINOVA_FILE = BASE_DIR / "bulletin_template_linova.html"


def adjust_grade(note):
    """Applique la formule f(x) = 0.57x + 8.74, plafonnee a 20
    Si pas de note, retourne une note aleatoire entre 9.5 et 12"""
    if note is None or note == "-" or note == "":
        return round(random.uniform(9.5, 12.0), 1)
    try:
        note_float = float(note)
        adjusted = 0.57 * note_float + 8.74
        return min(adjusted, 20.0)
    except (ValueError, TypeError):
        return round(random.uniform(9.5, 12.0), 1)


def parse_note_string(note_str):
    """Parse le format 'XX.XX / 20' vers un float"""
    if note_str is None or str(note_str).strip() == "":
        return None
    try:
        s = str(note_str).strip()
        if "/" in s:
            parts = s.split("/")
            return float(parts[0].strip())
        else:
            return float(s)
    except (ValueError, IndexError):
        return None


def get_appreciation(note, matiere_nom, appreciations_dict):
    """Genere une appreciation basee sur la note et la matiere"""
    if note is None:
        note = 10.5

    appreciations = appreciations_dict.get(matiere_nom, {})

    if note < 10:
        return appreciations.get("insuffisant", "Des efforts sont nécessaires pour progresser.")
    elif note < 12:
        return appreciations.get("passable", "Résultats encourageants. Continuez vos efforts.")
    elif note < 14:
        return appreciations.get("assez_bien", "Bon travail. Poursuivez dans cette voie.")
    elif note < 16:
        return appreciations.get("bien", "Très bon travail. Continuez ainsi.")
    else:
        return appreciations.get("excellent", "Excellents résultats. Félicitations.")


def get_appreciation_generale(prenom, moyenne):
    """Genere une appreciation generale personnalisee"""
    if moyenne is None:
        moyenne = 10.5

    if moyenne < 10:
        return f"{prenom} doit fournir davantage d'efforts pour progresser. Un travail régulier et soutenu permettra d'améliorer les résultats l'année prochaine."
    elif moyenne < 12:
        return f"{prenom} montre des résultats encourageants. Avec plus de régularité dans le travail, les résultats continueront de s'améliorer."
    elif moyenne < 14:
        return f"{prenom} fournit un bon travail cette année. Les bases sont acquises et les efforts doivent être maintenus pour progresser davantage."
    elif moyenne < 16:
        return f"{prenom} a fourni un travail régulier et rigoureux tout au long de l'année. Les résultats sont très satisfaisants. Continuez ainsi."
    else:
        return f"{prenom} a fourni un travail remarquable tout au long de l'année. Félicitations pour ces excellents résultats."


def format_note(note):
    """Formate une note pour l'affichage"""
    if note is None:
        return "NN"
    return f"{note:.1f}".replace(".", ",")


def format_date(date_value):
    """Formate une date pour l'affichage"""
    if date_value is None:
        return ""
    if hasattr(date_value, 'strftime'):
        return date_value.strftime("%d/%m/%Y")
    return str(date_value)


def capitalize_name(name):
    """Met la premiere lettre de chaque partie du nom en majuscule,
    en preservant les particules et tirets"""
    if not name:
        return ""
    # Traiter chaque partie separee par un espace
    parts = name.strip().split()
    result = []
    for part in parts:
        # Gerer les tirets composes (ex: Jean-Baptiste)
        if '-' in part:
            sub = part.split('-')
            result.append('-'.join(s[0].upper() + s[1:].lower() if s else s for s in sub))
        else:
            result.append(part[0].upper() + part[1:].lower() if part else part)
    return ' '.join(result)


def sanitize_filename(text):
    """Nettoie un texte pour en faire un nom de fichier valide"""
    text = unicodedata.normalize('NFKD', str(text))
    text = text.encode('ASCII', 'ignore').decode('ASCII')
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[-\s]+', '_', text)
    return text.lower()


def load_excel_data():
    """Charge les donnees depuis les fichiers Excel du dossier notes/
    Chaque fichier = 1 matiere. Les eleves sont fusionnes par prenom+nom."""
    students_map = {}  # cle = (prenom_lower, nom_lower) -> student dict

    for matiere in MATIERES_PAES:
        excel_file = matiere.get("excel_file")
        if not excel_file:
            continue

        filepath = NOTES_DIR / excel_file
        if not filepath.exists():
            print(f"  [WARN] Fichier non trouvé: {filepath}")
            continue

        wb = load_workbook(filepath)
        ws = wb.active

        # Lire les en-tetes pour identifier les colonnes
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value or "").strip())

        # Trouver les index des colonnes
        def find_col(names):
            for name in names:
                for i, h in enumerate(headers):
                    if h.lower() == name.lower():
                        return i
            return None

        col_prenom = find_col(["Prenom", "Prénom"])
        col_nom = find_col(["Nom"])
        col_email = find_col(["Email", "Mail"])
        col_note = find_col(["Note"])

        if col_prenom is None or col_nom is None:
            print(f"  [WARN] Colonnes Prenom/Nom non trouvées dans {excel_file}: {headers}")
            continue

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[col_nom] is None:
                continue

            prenom = capitalize_name(str(row[col_prenom] or ""))
            nom = capitalize_name(str(row[col_nom] or ""))
            email = str(row[col_email] or "").strip() if col_email is not None else ""

            if not prenom and not nom:
                continue

            key = (prenom.lower(), nom.lower())

            if key not in students_map:
                students_map[key] = {
                    "nom": nom,
                    "prenom": prenom,
                    "date_naissance": None,
                    "email": email,
                    "notes": {}
                }
            elif email and not students_map[key]["email"]:
                students_map[key]["email"] = email

            # Parser la note
            note_raw = row[col_note] if col_note is not None else None
            note_value = parse_note_string(note_raw)
            students_map[key]["notes"][matiere["nom"]] = note_value

        wb.close()

    # Charger les dates de naissance depuis Identites.xlsx
    identites_file = BASE_DIR / "Identites.xlsx"
    if identites_file.exists():
        print(f"  [OK] Chargement des identités depuis Identites.xlsx...")
        wb_id = load_workbook(identites_file)
        ws_id = wb_id.active
        matched = 0

        for row in ws_id.iter_rows(min_row=2, values_only=True):
            if row[2] is None or row[3] is None:
                continue
            nom_id = capitalize_name(str(row[2] or ""))
            prenom_id = capitalize_name(str(row[3] or ""))
            date_naiss = row[4]  # datetime object
            key = (prenom_id.lower(), nom_id.lower())

            if key in students_map:
                students_map[key]["date_naissance"] = date_naiss
                matched += 1

        wb_id.close()
        print(f"  [OK] {matched} dates de naissance associées")
    else:
        print(f"  [WARN] Fichier Identites.xlsx non trouvé")

    return list(students_map.values())


def load_identity_choices():
    """Charge Identites.xlsx et retourne un dict pour enrichir l'index.
    Retourne: {(prenom_lower, nom_lower): {"date_naissance": ..., "choix": ...}}
    choix = "Bulletin PAES" | "Bulletin Linova"
    """
    for filename in ("identity.xlsx", "Identites.xlsx"):
        filepath = BASE_DIR / filename
        if not filepath.exists():
            continue
        wb = load_workbook(filepath)
        ws = wb.active
        out = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) <= 3 or (row[2] is None and row[3] is None):
                continue
            nom = capitalize_name(str(row[2] or ""))
            prenom = capitalize_name(str(row[3] or ""))
            if not nom and not prenom:
                continue
            key = (prenom.lower(), nom.lower())
            date_naiss = row[4] if len(row) > 4 else None
            parcoursup = str(row[5] or "").strip() if len(row) > 5 else ""
            if "PAES" in parcoursup or "Diploma" in parcoursup:
                choix = "Bulletin PAES"
            elif "Linova" in parcoursup or "BTS" in parcoursup:
                choix = "Bulletin Linova"
            else:
                choix = "Formulaire non rempli"
            out[key] = {"date_naissance": date_naiss, "choix": choix}
        wb.close()
        print(f"  [OK] Identités chargées depuis {filename}: {len(out)} élèves")
        return out
    print(f"  [WARN] Aucun fichier identité trouvé (identity.xlsx / Identites.xlsx)")
    return {}


def calculate_class_stats(students, matieres_config, notes_key_fn=None):
    """Calcule les statistiques de classe pour chaque matiere.
    notes_key_fn: fonction qui prend une matiere et retourne la cle dans student['notes']
    """
    if notes_key_fn is None:
        notes_key_fn = lambda m: m["nom"]

    stats = {}

    for matiere in matieres_config:
        note_key = notes_key_fn(matiere)
        notes = []

        for student in students:
            raw_note = student["notes"].get(note_key)
            adjusted = adjust_grade(raw_note)
            if adjusted is not None:
                notes.append(adjusted)

        if notes:
            stats[matiere["nom"]] = {
                "moyenne": sum(notes) / len(notes),
                "min": min(notes),
                "max": max(notes)
            }
        else:
            stats[matiere["nom"]] = {
                "moyenne": None,
                "min": None,
                "max": None
            }

    return stats


def calculate_class_stats_from_adjusted(students, matieres_config, notes_key_fn=None):
    """Calcule les statistiques de classe a partir des notes PRE-CALCULEES (adjusted_notes).
    notes_key_fn: fonction qui retourne la cle dans student['adjusted_notes']
    """
    if notes_key_fn is None:
        notes_key_fn = lambda m: m["nom"]

    stats = {}

    for matiere in matieres_config:
        note_key = notes_key_fn(matiere)
        notes = []

        for student in students:
            adj = student.get("adjusted_notes", {}).get(note_key)
            if adj is not None:
                notes.append(adj)

        if notes:
            stats[matiere["nom"]] = {
                "moyenne": sum(notes) / len(notes),
                "min": min(notes),
                "max": max(notes)
            }
        else:
            stats[matiere["nom"]] = {
                "moyenne": None,
                "min": None,
                "max": None
            }

    return stats


def generate_bulletin_html(student, class_stats, template, profil, matieres_config,
                           appreciations_dict, notes_key_fn=None):
    """Genere le HTML d'un bulletin pour un eleve"""
    if notes_key_fn is None:
        notes_key_fn = lambda m: m["nom"]

    html = template

    # Informations etablissement
    html = html.replace("{{NOM_ETABLISSEMENT}}", profil["nom"])
    html = html.replace("{{ADRESSE_ETABLISSEMENT}}", profil["adresse"])
    html = html.replace("{{CODE_POSTAL}}", profil["code_postal"])
    html = html.replace("{{VILLE}}", profil["ville"])
    html = html.replace("{{ANNEE_SCOLAIRE}}", profil["annee_scolaire"])
    html = html.replace("{{CLASSE}}", profil["classe"])
    html = html.replace("{{CHARGE_ETUDES}}", profil["charge_etudes"])
    html = html.replace("{{SEMESTRE}}", profil["semestre"])
    html = html.replace("{{LOGO_FILE}}", profil["logo"])
    html = html.replace("{{TAMPON_FILE}}", profil["tampon"])

    # Informations eleve
    html = html.replace("{{PRENOM_ELEVE}}", str(student["prenom"] or ""))
    html = html.replace("{{NOM_ELEVE}}", str(student["nom"] or ""))
    date_naiss = format_date(student.get("date_naissance"))
    if date_naiss:
        html = html.replace("{{DATE_NAISSANCE}}", date_naiss)
    else:
        # Pas de date de naissance : supprimer toute la ligne "Né le : ..."
        html = html.replace('<div><span class="label">Né le : </span><span class="value field">{{DATE_NAISSANCE}}</span></div>', '')

    # Notes et appreciations par matiere
    student_notes = []

    for i, matiere in enumerate(matieres_config, start=1):
        note_key = notes_key_fn(matiere)
        # Utiliser les notes pre-calculees si disponibles (pour coherence PAES/Linova)
        adjusted_note = student.get("adjusted_notes", {}).get(note_key)
        if adjusted_note is None:
            raw_note = student["notes"].get(note_key)
            adjusted_note = adjust_grade(raw_note)

        html = html.replace(f"{{{{MATIERE_{i}}}}}", matiere["nom"])
        html = html.replace(f"{{{{ENSEIGNANT_{i}}}}}", matiere["enseignant"])
        html = html.replace(f"{{{{MOY_ELEVE_{i}}}}}", format_note(adjusted_note))

        # Statistiques de classe
        stats = class_stats.get(matiere["nom"], {})
        html = html.replace(f"{{{{MOY_CLASSE_{i}}}}}", format_note(stats.get("moyenne")))
        html = html.replace(f"{{{{NOTE_MIN_{i}}}}}", format_note(stats.get("min")))
        html = html.replace(f"{{{{NOTE_MAX_{i}}}}}", format_note(stats.get("max")))

        # Appreciation specifique a la matiere
        html = html.replace(f"{{{{APPRECIATION_{i}}}}}",
                           get_appreciation(adjusted_note, matiere["nom"], appreciations_dict))

        if adjusted_note is not None:
            student_notes.append(adjusted_note)

    # Moyennes generales
    if student_notes:
        moyenne_eleve = sum(student_notes) / len(student_notes)
    else:
        moyenne_eleve = 10.5

    # Calculer moyenne generale de la classe
    all_class_moyennes = []
    for matiere in matieres_config:
        stats = class_stats.get(matiere["nom"], {})
        if stats.get("moyenne") is not None:
            all_class_moyennes.append(stats["moyenne"])

    if all_class_moyennes:
        moyenne_classe = sum(all_class_moyennes) / len(all_class_moyennes)
    else:
        moyenne_classe = None

    html = html.replace("{{MOYENNE_GENERALE_ELEVE}}", format_note(moyenne_eleve))
    html = html.replace("{{MOYENNE_GENERALE_CLASSE}}", format_note(moyenne_classe))

    # Absences et appreciation generale
    html = html.replace("{{ABSENCES}}", "RAS")
    html = html.replace("{{APPRECIATION_GENERALE}}",
                       get_appreciation_generale(student["prenom"], moyenne_eleve))

    return html


def generate_index_html(students, profil_paes, profil_linova):
    """Genere la page index.html avec la liste des eleves et acces aux deux types de bulletins"""

    sorted_students = sorted(students, key=lambda s: (s["nom"] or "", s["prenom"] or ""))

    rows = []
    for student in sorted_students:
        filename_base = f"bulletin_{sanitize_filename(student['prenom'])}_{sanitize_filename(student['nom'])}"
        pdf_paes = f"bulletins_pdf_paes/{filename_base}.pdf"
        pdf_linova = f"bulletins_pdf_linova/{filename_base}.pdf"

        date_naiss = format_date(student.get("date_naissance"))
        if not date_naiss:
            date_naiss_html = '<span class="no-date">Formulaire non rempli</span>'
        else:
            date_naiss_html = date_naiss

        choix = student.get("choix", "Formulaire non rempli")

        row = f"""
            <tr>
                <td>{student['prenom'] or ''}</td>
                <td>{student['nom'] or ''}</td>
                <td><a href="mailto:{student.get('email') or ''}">{student.get('email') or ''}</a></td>
                <td>{date_naiss_html}</td>
                <td class="choix-cell">{choix}</td>
                <td class="pdf-cell pdf-cell-paes">
                    <a href="{pdf_paes}" class="btn-pdf btn-paes" target="_blank" title="Bulletin PAES">
                        <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" fill="currentColor" viewBox="0 0 16 16">
                            <path d="M14 14V4.5L9.5 0H4a2 2 0 0 0-2 2v12a2 2 0 0 0 2 2h8a2 2 0 0 0 2-2zM9.5 3A1.5 1.5 0 0 0 11 4.5h2V14a1 1 0 0 1-1 1H4a1 1 0 0 1-1-1V2a1 1 0 0 1 1-1h5.5v2z"/>
                        </svg>
                        PAES
                    </a>
                    <a href="{pdf_paes}" class="btn-download" download="{filename_base}_paes.pdf" title="Télécharger PAES">
                        <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" fill="currentColor" viewBox="0 0 16 16">
                            <path d="M.5 9.9a.5.5 0 0 1 .5.5v2.5a1 1 0 0 0 1 1h12a1 1 0 0 0 1-1v-2.5a.5.5 0 0 1 1 0v2.5a2 2 0 0 1-2 2H2a2 2 0 0 1-2-2v-2.5a.5.5 0 0 1 .5-.5z"/>
                            <path d="M7.646 11.854a.5.5 0 0 0 .708 0l3-3a.5.5 0 0 0-.708-.708L8.5 10.293V1.5a.5.5 0 0 0-1 0v8.793L5.354 8.146a.5.5 0 1 0-.708.708l3 3z"/>
                        </svg>
                    </a>
                </td>
                <td class="pdf-cell pdf-cell-linova">
                    <a href="{pdf_linova}" class="btn-pdf btn-linova" target="_blank" title="Bulletin Linova">
                        <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" fill="currentColor" viewBox="0 0 16 16">
                            <path d="M14 14V4.5L9.5 0H4a2 2 0 0 0-2 2v12a2 2 0 0 0 2 2h8a2 2 0 0 0 2-2zM9.5 3A1.5 1.5 0 0 0 11 4.5h2V14a1 1 0 0 1-1 1H4a1 1 0 0 1-1-1V2a1 1 0 0 1 1-1h5.5v2z"/>
                        </svg>
                        Linova
                    </a>
                    <a href="{pdf_linova}" class="btn-download btn-dl-linova" download="{filename_base}_linova.pdf" title="Télécharger Linova">
                        <svg xmlns="http://www.w3.org/2000/svg" width="14" height="14" fill="currentColor" viewBox="0 0 16 16">
                            <path d="M.5 9.9a.5.5 0 0 1 .5.5v2.5a1 1 0 0 0 1 1h12a1 1 0 0 0 1-1v-2.5a.5.5 0 0 1 1 0v2.5a2 2 0 0 1-2 2H2a2 2 0 0 1-2-2v-2.5a.5.5 0 0 1 .5-.5z"/>
                            <path d="M7.646 11.854a.5.5 0 0 0 .708 0l3-3a.5.5 0 0 0-.708-.708L8.5 10.293V1.5a.5.5 0 0 0-1 0v8.793L5.354 8.146a.5.5 0 1 0-.708.708l3 3z"/>
                        </svg>
                    </a>
                </td>
            </tr>"""
        rows.append(row)

    nb_students = len(students)
    nb_matieres = len(MATIERES_PAES)

    index_html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Bulletins Scolaires - {profil_paes['nom']} / {profil_linova['nom']}</title>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Open+Sans:wght@400;600;700&display=swap');

        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: 'Open Sans', Arial, sans-serif;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
            padding: 20px;
        }}

        .container {{
            max-width: 1400px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 10px 40px rgba(0,0,0,0.2);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #2c3e50 0%, #3498db 100%);
            color: white;
            padding: 30px 40px;
            text-align: center;
        }}

        .header h1 {{
            font-size: 28px;
            margin-bottom: 10px;
        }}

        .header p {{
            font-size: 14px;
            opacity: 0.9;
        }}

        .stats {{
            display: flex;
            justify-content: center;
            gap: 40px;
            margin-top: 20px;
            padding-top: 20px;
            border-top: 1px solid rgba(255,255,255,0.2);
        }}

        .stat {{
            text-align: center;
        }}

        .stat-value {{
            font-size: 32px;
            font-weight: 700;
        }}

        .stat-label {{
            font-size: 12px;
            opacity: 0.8;
        }}

        .search-container {{
            padding: 20px 40px;
            background: #f8f9fa;
            border-bottom: 1px solid #eee;
        }}

        .search-input {{
            width: 100%;
            padding: 12px 20px;
            font-size: 14px;
            border: 2px solid #e0e0e0;
            border-radius: 8px;
            outline: none;
            transition: border-color 0.3s;
        }}

        .search-input:focus {{
            border-color: #3498db;
        }}

        .table-container {{
            padding: 20px 40px 40px;
            overflow-x: auto;
        }}

        table {{
            width: 100%;
            border-collapse: collapse;
        }}

        th {{
            background: #f8f9fa;
            padding: 15px 12px;
            text-align: left;
            font-weight: 600;
            color: #2c3e50;
            border-bottom: 2px solid #dee2e6;
            font-size: 13px;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        td {{
            padding: 15px 12px;
            border-bottom: 1px solid #eee;
            vertical-align: middle;
        }}

        tr:hover {{
            background: #f8f9fa;
        }}

        a {{
            color: #3498db;
            text-decoration: none;
        }}

        a:hover {{
            text-decoration: underline;
        }}

        .pdf-cell {{
            display: flex;
            gap: 6px;
            align-items: center;
        }}

        .btn-pdf, .btn-download {{
            display: inline-flex;
            align-items: center;
            gap: 4px;
            padding: 6px 10px;
            border-radius: 6px;
            font-size: 11px;
            font-weight: 600;
            text-decoration: none;
            transition: all 0.2s;
        }}

        .btn-paes {{
            background: #e74c3c;
            color: white;
        }}

        .btn-paes:hover {{
            background: #c0392b;
            text-decoration: none;
        }}

        .btn-linova {{
            background: #2980b9;
            color: white;
        }}

        .btn-linova:hover {{
            background: #1f6fa3;
            text-decoration: none;
        }}

        .btn-download {{
            background: #27ae60;
            color: white;
            padding: 6px;
        }}

        .btn-download:hover {{
            background: #1e8449;
            text-decoration: none;
        }}

        .btn-dl-linova {{
            background: #16a085;
        }}

        .btn-dl-linova:hover {{
            background: #0e7e68;
        }}

        .no-date {{
            color: #e74c3c;
            font-size: 10px;
            font-style: italic;
        }}

        .choix-cell {{
            font-size: 12px;
            font-weight: 500;
        }}

        .pdf-cell-paes {{
            min-width: 100px;
            white-space: nowrap;
        }}

        .pdf-cell-linova {{
            min-width: 100px;
            white-space: nowrap;
        }}

        .no-results {{
            text-align: center;
            padding: 40px;
            color: #666;
            display: none;
        }}

        @media (max-width: 768px) {{
            .container {{
                border-radius: 0;
            }}

            .header, .search-container, .table-container {{
                padding-left: 20px;
                padding-right: 20px;
            }}

            .stats {{
                flex-direction: column;
                gap: 15px;
            }}

            th, td {{
                padding: 10px 8px;
                font-size: 12px;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Bulletins Annuels</h1>
            <p>{profil_paes['nom']} / {profil_linova['nom']} - Année scolaire {profil_paes['annee_scolaire']} - Classe {profil_paes['classe']}</p>
            <div class="stats">
                <div class="stat">
                    <div class="stat-value">{nb_students}</div>
                    <div class="stat-label">Élèves</div>
                </div>
                <div class="stat">
                    <div class="stat-value">{nb_matieres}</div>
                    <div class="stat-label">Matières</div>
                </div>
            </div>
        </div>

        <div class="search-container">
            <input type="text" class="search-input" id="searchInput"
                   placeholder="Rechercher un élève par nom, prénom ou email...">
        </div>

        <div class="table-container">
            <table id="studentsTable">
                <thead>
                    <tr>
                        <th>Prénom</th>
                        <th>Nom</th>
                        <th>Email</th>
                        <th>Date de naissance</th>
                        <th>Choix</th>
                        <th>Bulletin PAES</th>
                        <th>Bulletin Linova</th>
                    </tr>
                </thead>
                <tbody>
                    {''.join(rows)}
                </tbody>
            </table>
            <div class="no-results" id="noResults">
                Aucun élève trouvé pour cette recherche.
            </div>
        </div>
    </div>

    <script>
        document.getElementById('searchInput').addEventListener('input', function() {{
            const searchTerm = this.value.toLowerCase();
            const rows = document.querySelectorAll('#studentsTable tbody tr');
            let visibleCount = 0;

            rows.forEach(row => {{
                const text = row.textContent.toLowerCase();
                const isVisible = text.includes(searchTerm);
                row.style.display = isVisible ? '' : 'none';
                if (isVisible) visibleCount++;
            }});

            document.getElementById('noResults').style.display =
                visibleCount === 0 ? 'block' : 'none';
        }});
    </script>
</body>
</html>"""

    return index_html


def main():
    print("=" * 60)
    print("GENERATEUR DE BULLETINS SCOLAIRES")
    print("PAES (Diploma Santé) + Linova Education")
    print("=" * 60)

    # Creer les dossiers de sortie
    html_paes_dir = BASE_DIR / "bulletins_html_paes"
    pdf_paes_dir = BASE_DIR / "bulletins_pdf_paes"
    html_linova_dir = BASE_DIR / "bulletins_html_linova"
    pdf_linova_dir = BASE_DIR / "bulletins_pdf_linova"

    for d in [html_paes_dir, pdf_paes_dir, html_linova_dir, pdf_linova_dir]:
        d.mkdir(exist_ok=True)

    print(f"\n[OK] Dossiers de sortie créés")

    # Charger les templates
    print(f"[...] Chargement des templates...")
    with open(TEMPLATE_FILE, 'r', encoding='utf-8') as f:
        template = f.read()
    with open(TEMPLATE_LINOVA_FILE, 'r', encoding='utf-8') as f:
        template_linova = f.read()
    print(f"[OK] Templates chargés: PAES + Linova (1 page)")

    # Charger les donnees Excel
    print(f"[...] Chargement des données Excel depuis {NOTES_DIR}...")
    students = load_excel_data()
    print(f"[OK] {len(students)} élèves chargés")

    if not students:
        print("[ERREUR] Aucun élève trouvé. Vérifiez les fichiers Excel dans le dossier notes/")
        return

    # Enrichir les élèves avec les données du formulaire identité (date + choix)
    print(f"[...] Chargement du formulaire identité...")
    identity_dict = load_identity_choices()
    enriched_students = []
    for s in students:
        key = ((s.get("prenom") or "").lower(), (s.get("nom") or "").lower())
        if key in identity_dict:
            enriched_students.append({
                **s,
                "date_naissance": identity_dict[key]["date_naissance"],
                "choix": identity_dict[key]["choix"],
            })
        else:
            enriched_students.append({
                **s,
                "choix": "Formulaire non rempli",
            })

    # Vérifier s'il faut nettoyer (première exécution avec --clean ou --force-clean)
    # ou si on reprend après un crash (bulletins existants gardés)
    import sys
    if "--clean" in sys.argv:
        print(f"[...] Suppression de tous les anciens bulletins (--clean)...")
        for d in [html_paes_dir, pdf_paes_dir, html_linova_dir, pdf_linova_dir]:
            if d.exists():
                for f in d.glob("*"):
                    if f.is_file():
                        f.unlink()
        print(f"[OK] Anciens bulletins supprimés")
    else:
        existing_paes = len(list(pdf_paes_dir.glob("*.pdf"))) if pdf_paes_dir.exists() else 0
        existing_linova = len(list(pdf_linova_dir.glob("*.pdf"))) if pdf_linova_dir.exists() else 0
        print(f"[INFO] Reprise : {existing_paes} PAES et {existing_linova} Linova déjà générés")

    # Pre-calculer les notes ajustees UNE SEULE FOIS par eleve/matiere PAES
    # pour que PAES et Linova aient exactement les memes moyennes
    # Seed fixe pour que les notes soient identiques si on relance le script
    print(f"[...] Pré-calcul des notes ajustées (identiques PAES/Linova)...")
    random.seed(42)
    for student in enriched_students:
        adjusted = {}
        for matiere in MATIERES_PAES:
            raw_note = student["notes"].get(matiere["nom"])
            adjusted[matiere["nom"]] = adjust_grade(raw_note)
        student["adjusted_notes"] = adjusted
    print(f"[OK] Notes ajustées pré-calculées")

    # Calculer les statistiques de classe pour PAES (avec notes pre-calculees)
    print(f"[...] Calcul des statistiques de classe PAES...")
    class_stats_paes = calculate_class_stats_from_adjusted(enriched_students, MATIERES_PAES)
    print(f"[OK] Statistiques PAES calculées")

    # Calculer les statistiques de classe pour Linova (memes notes via mapping source_paes)
    print(f"[...] Calcul des statistiques de classe Linova...")
    linova_key_fn = lambda m: m.get("source_paes", m["nom"])
    class_stats_linova = calculate_class_stats_from_adjusted(enriched_students, MATIERES_LINOVA, notes_key_fn=linova_key_fn)
    print(f"[OK] Statistiques Linova calculées")

    # Generer les bulletins
    print(f"\n[...] Génération des bulletins...")

    for i, student in enumerate(enriched_students, start=1):
        prenom = student['prenom'] or 'inconnu'
        nom = student['nom'] or 'inconnu'
        filename_base = f"bulletin_{sanitize_filename(prenom)}_{sanitize_filename(nom)}"

        # === Bulletin PAES ===
        html_path_paes = html_paes_dir / f"{filename_base}.html"
        pdf_path_paes = pdf_paes_dir / f"{filename_base}.pdf"
        if not pdf_path_paes.exists():
            html_paes = generate_bulletin_html(
                student, class_stats_paes, template,
                PROFIL_PAES, MATIERES_PAES, APPRECIATIONS_PAES
            )
            with open(html_path_paes, 'w', encoding='utf-8') as f:
                f.write(html_paes)
            HTML(string=html_paes, base_url=str(BASE_DIR)).write_pdf(pdf_path_paes)
            del html_paes
            gc.collect()

        # === Bulletin Linova (meme mise en page A4 que PAES) ===
        html_path_linova = html_linova_dir / f"{filename_base}.html"
        pdf_path_linova = pdf_linova_dir / f"{filename_base}.pdf"
        if not pdf_path_linova.exists():
            html_linova = generate_bulletin_html(
                student, class_stats_linova, template_linova,
                PROFIL_LINOVA, MATIERES_LINOVA, APPRECIATIONS_LINOVA,
                notes_key_fn=linova_key_fn
            )
            with open(html_path_linova, 'w', encoding='utf-8') as f:
                f.write(html_linova)
            HTML(string=html_linova, base_url=str(BASE_DIR)).write_pdf(pdf_path_linova)
            del html_linova
            gc.collect()

        print(f"  [{i}/{len(enriched_students)}] {prenom} {nom} OK")

    print(f"\n[OK] {len(enriched_students)} x 2 bulletins générés ({len(enriched_students)} PAES + {len(enriched_students)} Linova)")

    # Generer l'index HTML (tous les élèves des notes, enrichis avec choix + date)
    print(f"\n[...] Génération de la plateforme index.html...")
    index_html = generate_index_html(enriched_students, PROFIL_PAES, PROFIL_LINOVA)
    index_path = BASE_DIR / "index.html"

    with open(index_path, 'w', encoding='utf-8') as f:
        f.write(index_html)

    print(f"[OK] Plateforme créée: {index_path}")

    # Resume
    print("\n" + "=" * 60)
    print("RÉSUMÉ")
    print("=" * 60)
    print(f"  Bulletins PAES HTML:   {html_paes_dir}")
    print(f"  Bulletins PAES PDF:    {pdf_paes_dir}")
    print(f"  Bulletins Linova HTML: {html_linova_dir}")
    print(f"  Bulletins Linova PDF:  {pdf_linova_dir}")
    print(f"  Plateforme:            {index_path}")
    print(f"\nOuvrez index.html dans votre navigateur pour accéder aux bulletins.")
    print("=" * 60)


if __name__ == "__main__":
    main()
